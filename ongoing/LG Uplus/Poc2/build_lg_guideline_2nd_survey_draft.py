#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""엑셀 `가이드라인 2차.xlsx` → `가이드라인_2차-survey-draft.json` (Poc2 survey_draft 스키마).

probing_plan 매핑:
- root_note: 반복 금지 한 줄만(엑셀 비고·참고는 넣지 않음).
- root_checklist: 추가로 던질 질문/프로빙 지시(가설 검증·심화 질문 등). `?`로 이어진 후속 질문, 비고 중 확인 지시 줄 등.
- branches: 질문당 최소 10개. 각 항목은 { id, if, note, position, checklist[{id,text,position}], branches[] }.
  · if: 참가자가 실제로 말할 법한 아주 구체적인 예상 발화(한두 문장).
  · checklist[0].text: 그 발화에 맞춘 팔로업 질문.
- 가이드 맨 앞 줄이 전부 `[…]`인 경우(세그먼트·스크리닝): 그 줄은 question.content에 넣지 않고 branch.note에만 둔다.
  본문의 실제 질문은 모든 분기의 checklist[0].text로 통일하고, content는 짧은 전환 문장(SEGMENT_BRIDGE_CONTENT)만 둔다.
"""

from __future__ import annotations

import hashlib
import json
import re
import uuid
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
EXCEL_PATH = HERE / "가이드라인 2차.xlsx"
OUT_JSON = HERE / "가이드라인_2차-survey-draft.json"
SHEET_MAIN = "가이드라인_수정"
SHEET_DESIGN = "조사설계"

FW_SPACE = "\u3000"

# 선행 `[…]` 한 줄(세그먼트/스크리닝)만 있는 가이드: 참가자에게 읽는 content가 아니라 분기·모더 참고용.
SEGMENT_BRIDGE_CONTENT = (
    "앞서 나눈 이야기를 이어서, 같은 주제로 한 가지만 더 여쭤볼게요."
)


def strip_leading_segment_lines(text: str) -> tuple[str, str]:
    """가이드 맨 앞의 `전체 한 줄이 [대괄호]`인 줄만 연속 제거 → (세그먼트 문자열, 나머지 본문)."""
    lines = (text or "").splitlines()
    segment_parts: list[str] = []
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        if not ln:
            i += 1
            continue
        if re.fullmatch(r"\[[^\]]+\]", ln):
            segment_parts.append(ln)
            i += 1
            continue
        break
    remainder = "\n".join(lines[i:]).strip()
    segment_note = "\n".join(segment_parts).strip()
    return segment_note, remainder


def nid() -> str:
    return str(uuid.uuid4())


def split_compound_question(text: str) -> tuple[str, list[str], list[str]]:
    """`?`로 분리: 첫 질문은 content, 이후 `?`가 있는 조각은 추가 질문(probe), 없으면 정보성(info)."""
    text = (text or "").strip()
    if not text or "?" not in text:
        return text, [], []

    parts = re.split(r"\?\s+", text)
    if len(parts) == 1:
        return text, [], []

    main = parts[0].strip() + "?"
    probes: list[str] = []
    info_bits: list[str] = []

    def flush_segment(seg: str) -> None:
        seg = seg.strip()
        if not seg:
            return
        if "?" not in seg:
            info_bits.append(seg)
            return
        # `?\s+`는 문장 끝 `?`(뒤 공백 없음)에서 잘리지 않아 `??` 중복이 날 수 있음
        sub = re.split(r"\?\s+", seg)
        if len(sub) == 1:
            one = sub[0].strip()
            probes.append(one if one.endswith("?") else one + "?")
            return
        first = sub[0].strip()
        if not first.endswith("?"):
            first += "?"
        probes.append(first)
        for s in sub[1:]:
            flush_segment(s)

    for segment in parts[1:]:
        flush_segment(segment)

    return main, probes, info_bits


def refine_probes_and_info_from_newlines(
    probes: list[str], info_bits: list[str]
) -> tuple[list[str], list[str]]:
    """한 덩어리 probe에 줄바꿈으로 안내+질문이 붙은 경우: 앞줄은 정보, 마지막 질문만 probe."""
    out_probes: list[str] = []
    out_info = list(info_bits)
    for p in probes:
        if "\n" not in p:
            out_probes.append(p)
            continue
        lines = [ln.strip() for ln in p.splitlines() if ln.strip()]
        if len(lines) >= 2 and lines[-1].endswith("?"):
            out_info.extend(lines[:-1])
            out_probes.append(lines[-1])
        else:
            out_probes.append(p)
    return out_probes, out_info


def is_definition_bullet(line: str) -> bool:
    return bool(re.match(r"^-\s*[^:]+:\s*.+", line.strip()))


def is_probe_note_line(line: str) -> bool:
    """비고 한 줄: 추가로 던져야 할 질문/프로빙 지시면 True, [참고]·[가설]·정의형 불릿은 False."""
    t = line.strip()
    if not t:
        return False
    if t.startswith("[참고]") or t.startswith("[가설]"):
        return False
    if t.startswith("▶"):
        return False
    if len(t) > 200:
        return False
    if t.endswith("?"):
        return True
    compact = t.replace(" ", "")
    if "확인필요" in compact:
        return True
    if is_definition_bullet(t):
        return False
    if t.startswith("-") and ("확인" in t or "파악" in t):
        return True
    return False


def split_note_into_info_and_probes(note) -> tuple[list[str], list[str]]:
    """비고 열: 정보성(참고·가설·범주 나열 등) vs 추가 질문용 프로빙."""
    raw_lines = note_lines(note)
    info: list[str] = []
    probes: list[str] = []
    for line in raw_lines:
        if is_probe_note_line(line):
            probes.append(line.strip())
        else:
            info.append(line.strip())
    return info, probes


ROOT_NOTE_SINGLE = "동일하거나 유사한 취지의 질문은 반복하지 않는다"


def build_root_note(_info_fragments: list[str]) -> str:
    """프로덕트 매핑용: 반복 금지 문장만 유지(정보·참고는 root_note에 넣지 않음)."""
    return ROOT_NOTE_SINGLE


def note_lines(note) -> list[str]:
    if note is None:
        return []
    s = str(note).strip()
    if not s or s == FW_SPACE:
        return []
    out: list[str] = []
    for line in s.splitlines():
        t = line.strip()
        if t:
            out.append(t)
    return out


def make_question(
    content: str,
    readable_id: str,
    excel_q: str,
    root_note: str,
    checklist: list[str],
    branches: list[dict],
    probing_guide: str | None,
) -> dict:
    pp = {"root_note": root_note, "root_checklist": [], "branches": branches or []}
    for pos, text in enumerate(checklist):
        pp["root_checklist"].append({"id": nid(), "text": text, "position": pos})
    return {
        "id": nid(),
        "readable_id": readable_id,
        "content": content.strip(),
        "question_type": "open_text",
        "options": None,
        "rating_min": None,
        "rating_max": None,
        "probing_guide": probing_guide,
        "probing_plan": pp,
        "response_type": "episode",
        "duration_sec": 45,
        "media": [],
    }


MIN_FOLLOW_UP_BRANCHES = 10

# (구체적 예상 발화, 그에 맞는 팔로업) — if 는 참가자가 실제로 말할 법한 문장


def topic_short(content: str, max_len: int = 44) -> str:
    s = re.sub(r"\[[^\]]*\]\s*", "", content or "")
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s or "이번 질문"


def _t(s: str, topic: str) -> str:
    return s.replace("{topic}", topic)


_JOIN_CHANNEL_PAIRS: list[tuple[str, str]] = [
    (
        "유플러스 직영 매장에서 가입했어요.",
        "그날 상담은 대략 몇 분 정도 걸렸고, 직원분이 특히 강조해서 설명해 주신 부분이 있었나요?",
    ),
    (
        "대리점에서 했습니다. 한 통신사만 파는 곳이었어요.",
        "대리점이라고 하셨는데, 다른 통신사 상품이랑 비교 견적도 같이 받아보셨나요, 아니면 한 곳만 보셨나요?",
    ),
    (
        "LG 유플러스 닷컴에서 혼자 온라인으로 가입했어요.",
        "온라인으로 진행하실 때 본인인증이나 서류 제출 중에 막히거나 헷갈린 단계가 있었나요?",
    ),
    (
        "114에 전화해서 상담 받고 가입 연결됐어요.",
        "전화 상담에서 들었던 월 요금이나 혜택 설명이, 개통 뒤 앱이나 문자로 본 내용이랑 다른 느낌은 없었나요?",
    ),
    (
        "지인이 다니는 판매점 소개받아서 그쪽으로 갔어요.",
        "소개로 가셨다고 하셨는데, 그날 계약서나 약정 기간에 대해 충분히 설명받았다고 느끼셨나요?",
    ),
    (
        "인터넷 설치 기사님 오실 때 같이 모바일도 개통해줬어요.",
        "현장에서 추가로 권유받은 부가서비스나 약정이 있었는지, 있다면 어떤 내용이었나요?",
    ),
    (
        "백화점 지하 통신 매장에서 했어요.",
        "그날 여러 통신사 부스를 돌아보셨나요, 아니면 한 군데만 보고 결정하셨나요?",
    ),
    (
        "가입은 빨리 끝났는데 개통 문자만 와서 자세히는 잘 몰라요.",
        "개통 문자 받으신 뒤에 앱 깔거나 설정하시면서 헷갈리신 적은 없었나요?",
    ),
    (
        "T월드 매장에서 SKT로 했어요.",
        "SKT 매장이라고 하셨는데, 번호이동 할인이나 기기 할부 조건은 어떻게 안내받으셨나요?",
    ),
    (
        "KT 공식 홈페이지에서 신청했습니다.",
        "신청 과정에서 자동으로 체크된 항목이나, 나중에 보니 의외였던 항목이 있었나요?",
    ),
]

_NPS_SCORE_PAIRS: list[tuple[str, str]] = [
    (
        "5점 정도 줄 것 같아요.",
        "10점 만점에 5점이라고 하셨는데, 한 점만 올리려면 통신사 쪽에서 무엇이 바뀌었으면 좋겠나요?",
    ),
    (
        "8점이요. 괜찮은 편이에요.",
        "8점과 9점 사이에서 고르라면 지금은 어느 쪽에 더 가깝게 느끼시나요, 그 이유는요?",
    ),
    (
        "추천은 잘 못 할 것 같아요.",
        "추천하기 어렵다고 느끼신 건 요금 때문인지, 품질 때문인지, 아니면 다른 이유인지 짧게 짚어 주실 수 있을까요?",
    ),
    (
        "0점이요. 전혀 추천 안 해요.",
        "그렇게까지 낮게 보시게 된 사건이나 순간이 가장 먼저 떠오르는 게 있나요?",
    ),
    (
        "10점 만점에 10점이요, 만족해요.",
        "그렇게 좋게 보시게 된 결정적인 경험이 있다면 한 가지만 구체적으로 말씀해 주실 수 있을까요?",
    ),
    (
        "한 6점? 아직 써 본 지 얼마 안 돼서 잘 모르겠어요.",
        "가입하신 지 대략 몇 주·몇 달 정도 되셨고, 그동안 가장 인상 깊었던 일이 있었나요?",
    ),
    (
        "강하게 추천할 것 같아요, 주변에 이미 말하기도 했어요.",
        "주변에 권유하실 때 특히 어떤 점을 강조해서 이야기하셨나요?",
    ),
    (
        "7점? 딱히 추천할 정도는 아닌 것 같아요.",
        "추천하기 애매하다고 느끼시는 부분이 요금인지, 서비스인지, 아니면 다른 이유인지 말씀해 주실 수 있을까요?",
    ),
    (
        "아직 잘 모르겠어요, 조금 더 써봐야 판단할 것 같아요.",
        "어느 정도 기간 더 써봐야 추천 여부를 결정할 수 있을 것 같으신가요?",
    ),
    (
        "3점이요. 솔직히 별로예요.",
        "3점이라고 하셨는데, 추천을 망설이게 되는 가장 큰 이유 한 가지만 꼽아 주신다면요?",
    ),
]

_REASON_PAIRS: list[tuple[str, str]] = [
    (
        "요금이 생각보다 많이 나와서요.",
        "가입 때 들었던 금액이랑 첫 청구서 금액이 달라 보였던 부분이 있었나요?",
    ),
    (
        "멤버십 혜택이 앱에서 잘 안 보여요.",
        "혜택을 쓰려고 시도해 보셨을 때 어느 단계에서 막히셨는지 말씀해 주실 수 있을까요?",
    ),
    (
        "이전 통신사랑 딱히 차이는 모르겠어요.",
        "통화 품질, 데이터 속도, 고객센터 중에서 ‘비슷하다’고 느끼신 영역이 어디에 가깝나요?",
    ),
    (
        "부가서비스 가입된 줄도 몰랐어요.",
        "언제·어떤 경로로 가입된 건지 나중에 어떻게 알게 되셨는지 순서대로 말씀해 주세요.",
    ),
    (
        "고객센터 전화가 너무 오래 걸려요.",
        "가장 최근에 연락하셨을 때 대기 시간이 대략 얼마나 됐고, 그때 문의하신 내용은 무엇이었나요?",
    ),
    (
        "약정 기간이랑 위약금 설명이 헷갈렸어요.",
        "설명은 어디서(매장·전화·문자 등) 받으셨고, 지금 생각하시면 어떤 문장이 애매했다고 느끼시나요?",
    ),
    (
        "앱이 복잡해서 찾기 힘들어요.",
        "특히 찾기 어려웠던 메뉴나 기능 이름이 기억나시는 대로 말씀해 주세요.",
    ),
    (
        "선물 받은 데이터가 바로 안 들어와서요.",
        "언제부터 들어올 거라고 안내받으셨고, 실제로는 언제쯤 반영됐나요?",
    ),
    (
        "스팸 문자가 많이 와서요.",
        "차단 설정을 해 보셨는지, 통신사 앱이나 부가서비스로 조치해 보신 적은 있나요?",
    ),
    (
        "특별히 이유는 없는데 그냥 별로예요.",
        "‘별로’라고 느끼신 감정이 가장 컸던 순간을 한 번만 떠올려 말씀해 주실 수 있을까요?",
    ),
]

_PLAN_ADDON_PAIRS: list[tuple[str, str]] = [
    (
        "OTT 포함된 7만 원대 요금제 썼어요.",
        "OTT는 실제로 얼마나 자주 보시는지, 요금제에 포함된 게 아깝다고 느끼신 적은 있나요?",
    ),
    (
        "인터넷이랑 휴대폰이랑 결합했어요.",
        "결합 할인 금액이 청구서에서 어떻게 표시되는지 확인해 보셨나요?",
    ),
    (
        "가족 네 줄 묶어서 할인 받는 걸로 했습니다.",
        "가족 결합 신청할 때 본인 말고 가족 분 동의나 확인 절차는 어떻게 진행됐나요?",
    ),
    (
        "V컬러링이랑 스팸차단 들어가 있어요.",
        "유료 부가서비스 중에 본인이 직접 고른 것과 자동으로 들어간 것이 있나요?",
    ),
    (
        "휴대폰 보험 같은 거는 안 넣었어요.",
        "안 넣으신 이유가 가격 때문이었는지, 필요 없다고 느끼셔서였는지 말씀해 주세요.",
    ),
    (
        "처음엔 무료라고 해서 넣었는데 나중에 유료로 바뀐 게 있어요.",
        "무료 기간이 끝난 뒤에 문자나 앱으로 따로 안내를 받으셨나요?",
    ),
    (
        "듀얼넘버 쓰고 있는데 잘 안 써요.",
        "가입할 때 실사용할 거라고 생각하셨는지, 아니면 권유로 들어가신 건지 기억나시나요?",
    ),
    (
        "요금제 바꾸려고 앱 봤는데 너무 어려워서 포기했어요.",
        "어느 화면까지 가셨다가 막히셨는지, 기억나는 메뉴 이름이 있을까요?",
    ),
    (
        "대리점에서 추천해준 대로 다 했어요.",
        "나중에 보니 본인이 원하지 않았던 항목이 있었는지 확인해 보신 적 있나요?",
    ),
    (
        "OTT는 거의 안 보고 데이터만 써요.",
        "그렇다면 지금 요금제가 본인 사용 패턴에 맞는지, 아쉬운 점이 있으신가요?",
    ),
]

_BILLING_PAIRS: list[tuple[str, str]] = [
    (
        "첫 달 청구서 보고 깜짝 놀랐어요.",
        "어떤 항목이 예상보다 컸는지, 청구서 화면에서 어떻게 표기돼 있었는지 말씀해 주세요.",
    ),
    (
        "할인이 다 반영된 건지 잘 모르겠더라고요.",
        "할인 내역을 확인하려고 고객센터나 앱을 찾아보셨나요?",
    ),
    (
        "자동이체 날짜만 알고 금액은 잘 몰랐어요.",
        "납부 전에 예상 금액을 확인할 수 있는 방법을 안내받으신 적이 있나요?",
    ),
    (
        "부가서비스 요금이 따로 붙어 있었어요.",
        "그 항목을 처음 본 게 청구서였는지, 미리 앱에서 본 적이 있는지 어느 쪽에 가깝나요?",
    ),
    (
        "가입 때 말한 금액이랑 달라요.",
        "가입 때는 얼마라고 들으셨고, 청구서에는 얼마로 나왔는지 대략 말씀해 주실 수 있을까요?",
    ),
    (
        "청구서는 봤는데 잘 이해는 못 했어요.",
        "특히 어떤 용어나 항목이 가장 헷갈리셨나요?",
    ),
    (
        "아직 청구서는 안 받아봤어요.",
        "그렇다면 첫 청구가 언제쯤일 거라고 안내받으셨는지 기억나시나요?",
    ),
    (
        "납부 알림 문자는 왔어요.",
        "문자에 적힌 금액이 본인이 예상했던 금액이랑 비슷했나요?",
    ),
    (
        "세금이나 부가세 때문에 더 나온 줄 알았는데 아니었어요.",
        "실제로는 어떤 항목 때문에 차이가 났는지 나중에 파악하셨나요?",
    ),
    (
        "전 달이랑 비교해 봤는데 요금이 올랐어요.",
        "요금이 오른 이유를 앱이나 상담으로 확인해 보신 적 있나요?",
    ),
]

_MEMBERSHIP_PAIRS: list[tuple[str, str]] = [
    (
        "스타벅스 쿠폰 자주 써요.",
        "쿠폰 받을 때 앱에서 인증이나 발급 과정에서 불편했던 점이 있었나요?",
    ),
    (
        "멤버십 있는 줄도 처음엔 몰랐어요.",
        "나중에 어떻게 알게 되셨고, 그때 바로 써 보셨나요?",
    ),
    (
        "포인트가 소멸됐다고 해서 화났어요.",
        "소멸 전에 통신사에서 알림을 받으셨는지, 어떤 채널로 받으셨나요?",
    ),
    (
        "제휴처가 마음에 안 들어요.",
        "본인이 자주 가는 브랜드가 포함돼 있지 않다고 느끼신 건가요?",
    ),
    (
        "멤버십 혜택이 예전보다 줄었어요.",
        "언제쯤부터 그렇게 느끼셨는지, 특정 공지나 이벤트가 있었나요?",
    ),
    (
        "가족이랑 같이 쓰는 혜택이 있으면 좋겠어요.",
        "지금은 본인만 쓰시는 구조인가요, 가족 카드나 계정 연동은 해 보셨나요?",
    ),
    (
        "쿠폰 고르는 화면이 너무 복잡해요.",
        "원하는 쿠폰 찾는 데 보통 몇 분 정도 걸리시는 편인가요?",
    ),
    (
        "실제로 한 달에 두세 번은 씁니다.",
        "혜택 쓸 때마다 만족도가 비슷한지, 가끔 실패하는 경우가 있는지 말씀해 주세요.",
    ),
    (
        "다른 카드 할인이 더 나아서 잘 안 써요.",
        "그래도 통신 멤버십을 유지하시는 이유가 따로 있으신가요?",
    ),
    (
        "적립은 되는데 쓸 데가 없어요.",
        "적립 내역은 앱에서 쉽게 확인되시나요?",
    ),
]

_EPISODE_PAIRS: list[tuple[str, str]] = [
    (
        "개통 첫날부터 데이터가 끊겼어요.",
        "끊긴 지역이 실내 특정 장소였는지, 이동 중이었는지 기억나시나요?",
    ),
    (
        "상담원이랑 말이 계속 엇갈렸어요.",
        "그 통화가 몇 차례 이어졌는지, 같은 상담원이었는지도 기억나시나요?",
    ),
    (
        "매장에서 기다리는 시간이 너무 길었어요.",
        "예약이나 순번표는 없었나요, 그냥 현장 대기였나요?",
    ),
    (
        "문자로 온 링크 눌렀더니 이상한 페이지로 갔어요.",
        "그 문자가 통신사 공식 번호에서 온 건지, 스팸 같았는지 어떻게 판단하셨나요?",
    ),
    (
        "교체 기기 받는 데 일주일 넘게 걸렸어요.",
        "그동안 임시 조치나 대체 단말 안내를 받으셨나요?",
    ),
    (
        "요금제 변경 신청했는데 반영이 안 됐어요.",
        "신청은 앱·전화·매장 중 어디로 하셨고, 처리 완료 문자는 받으셨나요?",
    ),
    (
        "다른 사람 명의로 잘못 개통될 뻔했어요.",
        "어떻게 알아채셨고, 이후에 통신사에서는 어떻게 조치했나요?",
    ),
    (
        "할부금이 청구서에 두 번 나온 것 같아요.",
        "고객센터에 문의해 보셨을 때 답변은 어떻게 나왔나요?",
    ),
    (
        "불편했지만 그냥 넘어갔어요.",
        "넘어가기로 하신 이유가 시간이 없어서였는지, 문의 방법을 몰라서였는지요?",
    ),
    (
        "지금은 괜찮은데 첫 주에만 그랬어요.",
        "첫 주 문제는 나중에 자연히 해결됐는지, 본인이 따로 조치하신 건지요?",
    ),
]

_YESNO_PAIRS: list[tuple[str, str]] = [
    (
        "네, 확인했어요.",
        "어떤 경로로 확인하셨고, 확인하실 때 가장 먼저 본 항목은 무엇이었나요?",
    ),
    (
        "아니요, 아직 안 봤어요.",
        "안 보신 이유가 바빠서인지, 어디서 봐야 할지 몰라서인지 어느 쪽에 가깝나요?",
    ),
    (
        "대충만 봤어요.",
        "자세히 안 보신 부분 중에 나중에 문제가 될까 걱정되는 게 있으신가요?",
    ),
    (
        "가입할 때 직원이 보여줬어요.",
        "그때 화면에서 본 금액이나 조건이, 나중에 본 것과 달랐던 적이 있나요?",
    ),
    (
        "앱 알림만 와서 그걸로 알았어요.",
        "알림만으로 충분하다고 느끼셨는지, 더 설명이 필요했다고 느끼셨는지요?",
    ),
    (
        "배우자가 대신 봤어요.",
        "본인이 직접 확인하지 않으신 이유가 있으셨나요?",
    ),
    (
        "몰라서 못 봤어요.",
        "지금 생각하면 어디를 열어보면 됐을 것 같다는 생각이 드시나요?",
    ),
    (
        "봤는데 이해가 안 돼서 그냥 뒀어요.",
        "가장 이해 안 됐던 문구나 숫자가 기억나시는 대로 말씀해 주세요.",
    ),
    (
        "네, 신청했어요.",
        "신청 과정에서 동의 화면이나 체크박스를 꼼꼼히 읽어 보셨나요?",
    ),
    (
        "아니요, 딱히 신청한 건 없어요.",
        "그런데도 서비스나 옵션이 들어가 있는 걸 나중에 발견하신 적은 있나요?",
    ),
]

_COMPARE_PAIRS: list[tuple[str, str]] = [
    (
        "전에 쓰던 SKT가 통화 품질은 더 나은 것 같아요.",
        "통화 품질 말고 데이터나 가격은 지금 회사가 더 나은 편인가요?",
    ),
    (
        "KT랑 별 차이 없어요.",
        "그럼에도 지금 회사를 쓰시는 가장 큰 이유는 무엇인가요?",
    ),
    (
        "LG유플러스가 요금은 저렴해서 옮겼어요.",
        "옮긴 뒤 요금은 기대에 맞았는지, 다른 데서 아쉬움이 생기진 않았나요?",
    ),
    (
        "이전 회사는 앱이 더 편했어요.",
        "지금 앱에서 특히 불편한 화면이나 기능이 있다면 한 가지 꼽아 주실 수 있을까요?",
    ),
    (
        "고객센터는 예전 게 더 빨랐어요.",
        "최근에 지금 통신사 고객센터 써 보신 경험이 있으신가요?",
    ),
    (
        "번호이동 할인 때문에 갈아탔어요.",
        "할인 기간이 끝난 뒤 요금이 어떻게 바뀔지 안내받으셨나요?",
    ),
    (
        "가족이 다 여기 써서 따라왔어요.",
        "본인이 직접 고른 게 아니라면, 써 보신 뒤 만족도는 어떠신가요?",
    ),
    (
        "딱히 불만은 없는데 예전이 더 익숙했어요.",
        "익숙해지기까지 앱이나 설정에서 어려우셨던 부분이 있었나요?",
    ),
    (
        "데이터 속도는 지금이 더 빨라요.",
        "속도는 만족인데 다른 영역에서 트레이드오프가 있다고 느끼시나요?",
    ),
    (
        "혜택 구성은 예전 게 더 마음에 들었어요.",
        "지금 멤버십에서 가장 아쉬운 점 한 가지를 말씀해 주실 수 있을까요?",
    ),
]

_SECURITY_PAIRS: list[tuple[str, str]] = [
    (
        "본인인증 문자가 너무 자주 와서 불안해요.",
        "최근에 받으신 인증 문자 중에 본인이 요청하지 않은 것 같은 게 있었나요?",
    ),
    (
        "제3자에게 개인정보 넘어갈까 걱정돼요.",
        "그 걱정이 특히 커지신 계기가 가입 과정이었는지, 이후 사용 중이었는지요?",
    ),
    (
        "상담원한테 주민번호 끝까지 말해야 해서 싫었어요.",
        "다른 확인 방법이 있다고 안내받으신 적은 있었나요?",
    ),
    (
        "스미싱 같은 거 올까 봐 링크 잘 안 눌러요.",
        "통신사 공식 문자와 광고 문자를 구분하실 때 기준으로 삼으시는 게 있나요?",
    ),
    (
        "앱 권한이 너무 많이 필요해서 찝찝했어요.",
        "그중에 거부하거나 나중에 끄신 권한이 있나요?",
    ),
    (
        "다른 사람 휴대폰으로 내 요금제가 보일까 봐요.",
        "실제로 그런 일이 있었는지, 아니면 걱정만 하신 건지요?",
    ),
    (
        "위치 정보 켜야 하는 게 싫었어요.",
        "끄면 서비스 제한된다는 안내를 보셨을 때 어떻게 하셨나요?",
    ),
    (
        "특별히 문제는 없었어요.",
        "그럼에도 불안하다고 느끼시는 순간이 있다면 언제인가요?",
    ),
    (
        "명의 도용될까 봐 가입할 때 조심했어요.",
        "조심하신다고 하셨을 때 구체적으로 어떤 점을 확인하셨나요?",
    ),
    (
        "통신사는 믿는데 앱 마케팅 동의가 번거로워요.",
        "동의 항목 중에 나중에 후회한 체크가 있으셨나요?",
    ),
]

_CHURN_PAIRS: list[tuple[str, str]] = [
    (
        "한번 해지 상담까지는 받아봤어요.",
        "그때 유지 혜택으로 제안받은 내용이 기억나시는 대로 말씀해 주세요.",
    ),
    (
        "친구가 그만두라고 해서 잠깐 생각은 했어요.",
        "그래도 유지하신 결정적 이유가 있으셨나요?",
    ),
    (
        "요금 올라가면 바로 해지할 생각이에요.",
        "‘요금 올라감’의 기준을 본인은 어느 정도로 두고 계신가요?",
    ),
    (
        "아직은 없었어요.",
        "그동안 불만이 생겨도 유지하셨던 이유는 무엇이라고 생각하시나요?",
    ),
    (
        "앱에서 해지 메뉴 찾다가 포기했어요.",
        "해지를 쉽게 못 찾게 해놨다고 느끼셨나요?",
    ),
    (
        "위약금 때문에 못 나가요.",
        "위약금 금액은 얼마라고 안내받으셨는지 대략 기억나시나요?",
    ),
    (
        "일단 써 보고 판단하려고요.",
        "몇 달 정도 써 보고 결정하실 계획이신가요?",
    ),
    (
        "가족이 결합돼 있어서 혼자 못 끊어요.",
        "결합 구조 때문에 불편하신 점이 따로 있으신가요?",
    ),
    (
        "해지하려다가 상담 잘해줘서 유지했어요.",
        "어떤 말이나 조건이 마음이 바뀌게 하셨나요?",
    ),
    (
        "지금 당장은 아니고 나중에 보려고요.",
        "‘나중’이 되면 가장 먼저 보실 지표는 무엇일 것 같으신가요?",
    ),
]

# --- 문자/알림 안내 (Q10~Q14) ---
_MSG_NOTIFICATION_PAIRS: list[tuple[str, str]] = [
    ("가입하고 나서 문자가 하루에 서너 개씩 왔어요.", "그 문자들 중에 직접 열어서 끝까지 읽은 게 있나요?"),
    ("카톡 알림으로 왔는데 다른 알림이랑 섞여서 잘 안 봤어요.", "카톡 말고 문자(SMS)로 왔으면 더 잘 보셨을 것 같으신가요?"),
    ("개통 확인 문자만 기억나고 나머지는 모르겠어요.", "혹시 요금 안내나 혜택 관련 문자도 왔는데 넘기신 걸 수도 있을까요?"),
    ("처음 이틀은 많이 왔는데 그 후론 안 와요.", "처음 이틀에 온 문자 중 기억나는 내용이 있으신가요?"),
    ("너무 많아서 스팸 같았어요.", "어떤 내용이 특히 스팸처럼 느껴지셨나요?"),
    ("오히려 안내가 너무 적어서 불안했어요.", "어떤 내용이 왔으면 좋겠다고 느끼셨나요?"),
    ("RCS로 온 건 이미지가 있어서 잘 봤어요.", "이미지가 있어서 좋았던 구체적인 내용이 기억나시나요?"),
    ("문자가 왔는데 링크를 누를지 말지 고민했어요.", "통신사 공식 문자라고 믿게 된 기준이 있었나요?"),
    ("멤버십이나 혜택 안내 문자가 제일 많았어요.", "그 혜택을 실제로 써 보신 적이 있으신가요?"),
    ("문자가 온 건 알겠는데 내용을 이해 못 했어요.", "어떤 단어나 표현이 이해가 안 됐는지 기억나시나요?"),
    ("알림톡이 편했어요, 링크 바로 눌렀거든요.", "링크 눌러서 들어간 페이지가 기대한 내용이었나요?"),
    ("문자 안내 타이밍이 늦어서 이미 고객센터에 전화했었어요.", "문자가 좀 더 일찍 왔으면 전화 안 하셨을 것 같으신가요?"),
]

# --- 타임라인/가입 직후 경험 (Q7~Q9) ---
_TIMELINE_PAIRS: list[tuple[str, str]] = [
    ("개통 당일에 유심 꽂고 바로 됐어요.", "유심 꽂은 뒤 설정하면서 헷갈리거나 추가 안내가 필요했던 건 없었나요?"),
    ("이틀째에 앱 깔라는 문자가 와서 깔았어요.", "앱 설치 후 처음 열었을 때 어떤 화면이 나왔는지 기억나시나요?"),
    ("사흘째에 이전 통신사 번호가 완전히 넘어왔어요.", "번호 이전되는 동안 전화나 문자가 안 되던 시간이 있었나요?"),
    ("일주일쯤 지나서 첫 혜택 문자를 받았어요.", "그 혜택을 실제로 쓰셨는지, 아니면 지나치셨는지 기억나시나요?"),
    ("한 2주쯤 됐을 때 청구서 알림이 왔어요.", "금액을 보셨을 때 예상했던 것과 비교해서 어떠셨나요?"),
    ("가입 당일에 직원이 앱 설정까지 다 해줬어요.", "나중에 혼자 앱을 열었을 때도 쓸 수 있었나요, 다시 봐야 했나요?"),
    ("개통 후 3일 정도 데이터가 느렸어요.", "느리다고 느낀 게 어떤 상황(동영상·지도·검색 등)에서였나요?"),
    ("첫 주에 결합 신청하려고 다시 매장에 갔어요.", "한 번에 안 되고 다시 가셨을 때 어떤 기분이셨나요?"),
    ("가입 후 아무 일도 안 일어나서 오히려 불안했어요.", "그 불안이 어디서 오는 건지, 뭔가 빠진 느낌이었는지 말씀해 주세요."),
    ("처음 며칠은 요금이 얼마나 나올지 걱정됐어요.", "가입 시 들은 예상 금액이 기억나시나요?"),
]

# --- 고객지원/문의채널 (Q15~Q17) ---
_SUPPORT_CHANNEL_PAIRS: list[tuple[str, str]] = [
    ("고객센터 전화가 너무 오래 걸려서 끊었어요.", "대기 중에 콜백 서비스나 채팅 상담 안내를 받으신 적이 있나요?"),
    ("앱 채팅으로 문의했는데 AI가 답해서 답답했어요.", "AI 답변 후에 실제 상담원으로 연결이 됐나요?"),
    ("매장에 다시 갔더니 담당자가 바뀌어서 처음부터 설명했어요.", "이전 상담 기록이 공유 안 됐다고 느끼신 건가요?"),
    ("114 말고 다른 번호에서도 전화가 와서 혼란스러웠어요.", "그 번호가 통신사 공식인지 아닌지 어떻게 판단하셨나요?"),
    ("카카오톡 상담이 있는 줄 몰랐어요.", "알았더라면 더 일찍 쓰셨을 것 같으신가요?"),
    ("매장 방문 예약이 안 돼서 그냥 가서 기다렸어요.", "대기 시간이 대략 얼마나 됐고, 그 시간 동안 어떤 기분이셨나요?"),
    ("상담원마다 말이 달라서 누구 말을 믿어야 할지 모르겠어요.", "가장 최근에 엇갈렸던 내용이 무엇이었는지 기억나시나요?"),
    ("앱에서 FAQ 검색했는데 원하는 답이 안 나왔어요.", "어떤 키워드로 검색하셨는지 대략 말씀해 주시겠어요?"),
    ("전화하면 자동응답(ARS)이 너무 복잡해요.", "원하는 메뉴까지 몇 단계 거치셨는지 기억나시나요?"),
    ("해결은 됐는데 상담 끝나고 만족도 묻는 문자는 안 왔어요.", "통신사가 사후 확인을 해줬으면 좋겠다고 느끼신 건가요?"),
]

# --- 요금제 선택 이유/사용패턴 (Q18~Q22) ---
_PLAN_CHOICE_PAIRS: list[tuple[str, str]] = [
    ("직원이 이 요금제가 제일 낫다고 추천해줬어요.", "본인이 직접 비교해 보신 다른 요금제가 있었나요?"),
    ("데이터 무제한이 필요해서 골랐어요.", "실제로 한 달에 데이터를 대략 얼마나 쓰시나요?"),
    ("OTT가 포함돼 있어서 골랐는데 아직 안 썼어요.", "OTT 설정을 안 하신 이유가 몰라서인지, 필요 없어서인지요?"),
    ("이전 통신사랑 비슷한 요금대로 맞췄어요.", "같은 가격에 혜택 구성은 이전이랑 비교해서 어떠신가요?"),
    ("사은품이랑 할인 조건 때문에 그 요금제로 했어요.", "사은품을 받으셨나요? 받았다면 사용하고 계신가요?"),
    ("3개월 의무 사용이라 좀 걱정되긴 해요.", "3개월 후에 어떤 요금제로 바꿀 생각이 있으신가요?"),
    ("사실 요금제 차이를 잘 모르겠어요.", "가입할 때 요금제 비교표 같은 걸 보신 적이 있나요?"),
    ("가족이 다 같은 요금제라서 맞췄어요.", "혼자만 다른 요금제 쓰면 결합 할인에 차이가 생기나요?"),
    ("월 7만 원대 요금제인데 좀 비싸다고 느껴요.", "비싸다고 느끼시는 금액 기준이 대략 얼마쯤인가요?"),
    ("인터넷 결합으로 할인받으니까 괜찮아요.", "결합 할인이 없었으면 이 요금제를 선택하셨을까요?"),
]

# --- 네트워크 품질 (Q47~Q50) ---
_NETWORK_PAIRS: list[tuple[str, str]] = [
    ("집에서 와이파이 끄면 데이터가 느려요.", "집 주소가 아파트인지 단독주택인지, 층수도 기억나시나요?"),
    ("지하철에서 영상이 자주 끊겨요.", "어떤 노선·구간에서 특히 그런지 기억나시나요?"),
    ("통화 중에 상대방 목소리가 끊겼어요.", "실내였는지 이동 중이었는지, 하루에 몇 번 정도 그런지요?"),
    ("개통 후 3일 정도 5G가 안 잡혔어요.", "LTE로는 잘 됐고 5G만 안 잡힌 건가요?"),
    ("이전 통신사보다 확실히 느린 것 같아요.", "속도 측정 앱 같은 걸로 비교해 보신 적이 있으신가요?"),
    ("특정 시간대에만 느려지더라고요.", "주로 몇 시쯤에 느려지시는지 기억나시나요?"),
    ("아직 한 번도 끊긴 적 없어요.", "혹시 주로 실내에서만 쓰시는 편인가요?"),
    ("통신사에 신고했더니 기지국 문제라고 하더라고요.", "그 이후로 개선이 되셨나요?"),
    ("와이파이콜링 쓰라고 했는데 설정이 어려웠어요.", "설정 가이드를 어디서(앱·문자·상담원) 받으셨나요?"),
    ("출장 갈 때 시골에서 아예 안 터졌어요.", "이전 통신사에서도 같은 지역에서 안 터졌었나요?"),
]

# --- 결합 (Q29~Q35) ---
_COMBINE_PAIRS: list[tuple[str, str]] = [
    ("인터넷이랑 TV 같이 결합했어요.", "결합 후 월 할인이 얼마나 되는지 확인해 보셨나요?"),
    ("가족 3명이랑 무선 결합했어요.", "가족 분들 동의 받는 과정이 어떠셨나요?"),
    ("결합하려고 했는데 인터넷 약정이 남아서 못 했어요.", "약정 만료 시점을 확인하셨나요?"),
    ("결합 할인 금액이 생각보다 작았어요.", "가입 때 들었던 할인 금액이랑 실제 청구서가 달랐나요?"),
    ("결합은 매장에서만 된다고 해서 다시 갔어요.", "앱이나 전화로도 되면 좋겠다고 느끼셨나요?"),
    ("WAP 인증이 뭔지 몰라서 가족한테 물어봤어요.", "인증 절차를 설명받으셨을 때 이해가 되셨나요?"),
    ("기사 방문 예약이 2주나 걸렸어요.", "그 기간 동안 인터넷 없이 어떻게 하셨나요?"),
    ("결합 됐는지 안 됐는지 확인이 안 돼요.", "앱에서 결합 상태를 찾아보신 적이 있으신가요?"),
    ("아직 결합 안 했는데 할 생각은 있어요.", "결합을 미루고 계신 가장 큰 이유가 무엇인가요?"),
    ("유선 기사님이 오셔서 그때 같이 결합 처리했어요.", "현장에서 설명 듣고 바로 동의하신 건가요?"),
]

# --- 부가서비스 (Q24~Q28) ---
_ADDON_SERVICE_PAIRS: list[tuple[str, str]] = [
    ("부가서비스가 뭐가 있는지 사실 잘 몰라요.", "가입할 때 부가서비스 목록을 따로 보여주거나 설명받으신 적이 있나요?"),
    ("스팸차단은 필요해서 직접 넣었어요.", "앱에서 직접 신청하셨나요, 매장에서 넣으셨나요?"),
    ("무료라고 해서 넣었는데 3개월 뒤에 유료가 됐어요.", "유료 전환 전에 알림을 받으셨나요?"),
    ("V컬러링이 돈 나가는 줄 몰랐어요.", "청구서 보고 알게 되셨나요, 다른 경로로 알게 되셨나요?"),
    ("통화중 대기가 안 돼서 불편했어요.", "이전 통신사에서는 기본으로 되던 건가요?"),
    ("직원이 넣어준 거라 뭐가 있는지도 몰라요.", "앱에서 현재 가입된 부가서비스 목록을 확인해 보신 적이 있나요?"),
    ("부가서비스 해지하려고 했는데 메뉴를 못 찾겠어요.", "어느 메뉴까지 들어가셨다가 막히셨나요?"),
    ("필요 없는 거 빼려고 전화했더니 유지해달라고 하더라고요.", "그 대화에서 어떤 혜택을 제안받으셨나요?"),
    ("듀얼넘버 쓸 일이 없어서 해지하고 싶어요.", "해지하면 바로 되는 건지, 위약금이 있는 건지 확인하셨나요?"),
    ("부가서비스 금액이 합쳐지면 꽤 나와요.", "대략 월 얼마 정도 부가서비스 비용이 나가시나요?"),
]


def _generic_concrete_pairs(topic: str) -> list[tuple[str, str]]:
    return [
        ("네, 그랬어요.", "방금 '그랬다'고 하신 부분을 시간 순서로 조금만 더 풀어서 말씀해 주실 수 있을까요?"),
        ("아니요, 없었어요.", "없었다고 하셨는데 혹시 잠깐이라도 불편했던 순간은 없었는지 떠올려보시겠어요?"),
        (
            "잘 기억은 안 나요.",
            f"괜찮습니다. 「{topic}」와 관련해 가장 최근에 떠오르는 장면 하나만 짧게 말씀해 주실 수 있을까요?",
        ),
    ]


def collect_concrete_if_follow_pairs(content: str, info_blob: str, topic: str) -> list[tuple[str, str]]:
    c = re.sub(r"\[[^\]]*\]\s*", "", content or "")
    c_flat = re.sub(r"\s+", " ", c).strip()
    blob = f"{content or ''}\n{info_blob or ''}"

    acc: list[tuple[str, str]] = []

    # 가입 채널
    if re.search(r"가입.*어디|어디서.*가입|어디\s*에서.*했", c_flat):
        acc.extend(_JOIN_CHANNEL_PAIRS)
        acc.extend(_generic_concrete_pairs(topic))
        return _dedup_and_pad(acc, topic)

    # NPS 점수 — 이유/계기가 핵심인 경우 제외 (배타적: 이 풀만 사용)
    if re.search(r"추천.*생각|몇\s*점|점\s*주|NPS|어느\s*정도\s*있", c_flat) and not re.search(
        r"이유|왜\s*그렇|평가한\s*이유", c_flat
    ):
        acc.extend(_NPS_SCORE_PAIRS)
        acc.extend(_generic_concrete_pairs(topic))
        return _dedup_and_pad(acc, topic)

    # 요금제 선택 이유 (Q18 류) — "이유" 보다 먼저 확인 (배타적)
    if re.search(r"요금제.*선택|선택한.*요금제|요금제를\s*선택|선택.*이유|사용\s*패턴|의무\s*사용|변경\s*가능|본인이\s*선택", c_flat):
        acc.extend(_PLAN_CHOICE_PAIRS)
        acc.extend(_generic_concrete_pairs(topic))
        return _dedup_and_pad(acc, topic)

    # 이유/계기 — 해지 계기는 churn도 함께 추가 (배타적)
    if re.search(r"이유|왜\s*그렇|계기|무엇이.*었", c_flat):
        acc.extend(_REASON_PAIRS)
        if "해지" in c_flat:
            acc.extend(_CHURN_PAIRS)
        acc.extend(_generic_concrete_pairs(topic))
        return _dedup_and_pad(acc, topic)

    # 결합 (c_flat 기준: 결합 자체가 질문 핵심인 경우)
    if re.search(r"결합", c_flat):
        acc.extend(_COMBINE_PAIRS)

    # 부가서비스 (c_flat 기준)
    if re.search(r"부가서비스|유료\s*부가|무료\s*부가|통화중\s*대기|듀얼넘버|V컬러링", c_flat):
        acc.extend(_ADDON_SERVICE_PAIRS)

    # 요금제/OTT (결합·부가서비스에 안 걸린 일반 요금 이야기)
    if re.search(r"요금제|OTT", blob) and not acc:
        acc.extend(_PLAN_ADDON_PAIRS)

    # 문자/알림 안내 (c_flat 기준: 문자 빈도·알림 내용이 질문 핵심인 경우만)
    if re.search(r"문자\s*\(|MMS|RCS|카톡\s*알림|알림톡|발송\s*빈도|가이드\s*안내|문자로\s*계속|알림.*받는것", c_flat):
        acc.extend(_MSG_NOTIFICATION_PAIRS)
    elif re.search(r"문자.*좋았|문자.*필요|알림\s*내용|문자.*어느\s*정도", c_flat):
        acc.extend(_MSG_NOTIFICATION_PAIRS)

    # 타임라인/가입 직후 과정
    if re.search(r"당일부터|하루\s*단위|이벤트.*순서|2주|타임라인|헷갈리거나|불안하거나|짜증", c_flat):
        acc.extend(_TIMELINE_PAIRS)

    # 고객지원/문의 채널 (c_flat 우선)
    if re.search(r"문의|고객지원|고객센터|이상적인.*고객|연락.*채널", c_flat):
        acc.extend(_SUPPORT_CHANNEL_PAIRS)

    # 청구서 (c_flat 기준)
    if re.search(r"청구|청구서|빌링|납부|첫\s*달\s*청구", c_flat):
        acc.extend(_BILLING_PAIRS)

    # 멤버십/장기고객 (c_flat 기준)
    if re.search(r"멤버십|장기\s*고객|등급.*승급|어떤\s*등급|혜택.*관심|선착순|사용.*적\s*있", c_flat):
        acc.extend(_MEMBERSHIP_PAIRS)

    # 네트워크 품질 (c_flat 기준)
    if re.search(r"품질|속도|끊김|네트워크|5G|점검.*가이드|셀프\s*점검|스스로\s*점검", c_flat):
        acc.extend(_NETWORK_PAIRS)

    # 불편 에피소드
    if re.search(r"불편|어떤\s*과정|어떤\s*일|대응|에피소드|자세하게", c_flat):
        acc.extend(_EPISODE_PAIRS)

    # 확인 여부
    if re.search(r"확인.*하셨|보셨나요|있으셨나요|간략하게\s*설명", c_flat):
        acc.extend(_YESNO_PAIRS)

    # 이전 통신사 비교 (비교/차이/대비가 질문 핵심일 때만)
    if re.search(r"이전.*통신사.*비교|타사.*대비|비교.*이전|차이.*어떻|어떤.*차이|타사\s*대비|이전.*통신사.*차이", c_flat):
        acc.extend(_COMPARE_PAIRS)
    elif re.search(r"이전.*통신사|타사|옮기", c_flat) and re.search(r"비교|차이|대비", c_flat):
        acc.extend(_COMPARE_PAIRS)

    # 보안/개인정보 (c_flat 기준)
    if re.search(r"개인정보|보안|스미싱", c_flat):
        acc.extend(_SECURITY_PAIRS)

    # 해지 (c_flat 기준)
    if "해지" in c_flat:
        acc.extend(_CHURN_PAIRS)

    # 행동 질문 (Q9 류)
    if re.search(r"취한\s*행동|본인이\s*취한|어떻게\s*했|행동은", c_flat):
        acc.extend(_EPISODE_PAIRS)
        acc.extend(_SUPPORT_CHANNEL_PAIRS)

    # 개선/제안 질문 (Q53 류)
    if re.search(r"고친다면|한\s*가지만|무엇을\s*고치", c_flat):
        acc.extend(_REASON_PAIRS)
        acc.extend(_SUPPORT_CHANNEL_PAIRS)

    # 범용은 마지막에 소량만 추가
    acc.extend(_generic_concrete_pairs(topic))

    return _dedup_and_pad(acc, topic)


def _dedup_and_pad(acc: list[tuple[str, str]], topic: str) -> list[tuple[str, str]]:
    seen: set[str] = set()
    out: list[tuple[str, str]] = []
    for iff, fu in acc:
        k = iff.strip()
        if k in seen:
            continue
        seen.add(k)
        out.append((iff, fu))
    j = 0
    while len(out) < MIN_FOLLOW_UP_BRANCHES:
        j += 1
        out.append(
            (
                f"말로 하려니 잘 정리가 안 돼요. ({j}) 그때 기분만 기억나요.",
                f"괜찮습니다. 「{topic}」에 대해 떠오르는 단어 세 개만 먼저 말씀해 주시겠어요?",
            )
        )
    return out



def build_follow_up_branches(
    content: str,
    info_blob: str,
    readable_id: str,
    min_count: int = MIN_FOLLOW_UP_BRANCHES,
    forced_followup_text: str | None = None,
    segment_mod_note: str = "",
    keyword_content: str | None = None,
) -> list[dict]:
    """예상 발화(if) + 맞춤 팔로업(checklist.text). info_blob: 엑셀 비고+본문 안내(키워드 분기용).

    forced_followup_text: 선행 `[세그먼트]` 제거 뒤 본문 질문을 모든 분기의 팔로업으로 고정할 때 사용.
    segment_mod_note: 각 branch.note에 넣는 모더용 세그먼트 설명(참가자 content에 넣지 않음).
    keyword_content: forced 모드에서도 키워드 매칭용 텍스트(없으면 content 사용).
    """
    kc = (keyword_content if keyword_content is not None else content) or ""
    blob = info_blob
    if segment_mod_note:
        blob = f"{segment_mod_note}\n{blob}".strip()

    topic = topic_short(kc)
    unique = collect_concrete_if_follow_pairs(kc, blob, topic)

    h = int(hashlib.md5(readable_id.encode("utf-8")).hexdigest(), 16)
    start = h % len(unique)
    rotated = unique[start:] + unique[:start]
    selected = rotated[:min_count]

    forced = (forced_followup_text or "").strip()

    branches: list[dict] = []
    for pos, (iff, fu) in enumerate(selected):
        text = forced if forced else fu
        branches.append(
            {
                "id": nid(),
                "if": iff,
                "note": segment_mod_note,
                "position": pos,
                "checklist": [{"id": nid(), "text": text, "position": 0}],
                "branches": [],
            }
        )
    return branches


def read_design_meta(wb: openpyxl.Workbook) -> dict:
    meta: dict[str, str] = {}
    if SHEET_DESIGN not in wb.sheetnames:
        return meta
    ws = wb[SHEET_DESIGN]
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        if len(row) < 3:
            i += 1
            continue
        label = row[1]
        val_c = row[2]
        val_d = row[3] if len(row) > 3 else None
        ls = str(label).strip() if label else ""
        if ls == "조사 목적" and val_c:
            meta["goal"] = str(val_c).strip()
            i += 1
            continue
        if ls == "조사 대상자":
            parts: list[str] = []
            if val_c:
                if val_d and str(val_d).strip():
                    parts.append(f"{str(val_c).strip()}: {str(val_d).strip()}")
                else:
                    parts.append(str(val_c).strip())
            i += 1
            while i < len(rows):
                r2 = rows[i]
                b2 = r2[1] if len(r2) > 1 else None
                if b2 is not None and str(b2).strip():
                    break
                c2 = r2[2] if len(r2) > 2 else None
                d2 = r2[3] if len(r2) > 3 else None
                if c2 is None and (d2 is None or not str(d2).strip()):
                    i += 1
                    continue
                if c2 and d2 and str(d2).strip():
                    parts.append(f"{str(c2).strip()}: {str(d2).strip()}")
                elif c2 and str(c2).strip():
                    parts.append(str(c2).strip())
                i += 1
            meta["target_audience"] = "\n".join(parts)
            continue
        i += 1
    return meta


def section_from_col_b(cell) -> str | None:
    if cell is None:
        return None
    s = str(cell).strip()
    if not s or s == FW_SPACE:
        return None
    return s


def build_from_excel() -> dict:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_MAIN]
    rows = list(ws.iter_rows(values_only=True))
    meta = read_design_meta(wb)

    interview_chapters: list = []
    ch_idx = 0
    q_in_chapter = 0
    current_title: str | None = None
    current_questions: list = []

    def flush_chapter():
        nonlocal current_questions, ch_idx, q_in_chapter
        if not current_questions:
            q_in_chapter = 0
            return
        ch_idx = max(ch_idx, 1)
        interview_chapters.append(
            {
                "id": nid(),
                "readable_id": f"chapter_{ch_idx}",
                "title": current_title or f"챕터 {ch_idx}",
                "topics": [
                    {
                        "id": nid(),
                        "readable_id": f"topic_{ch_idx}_1",
                        "title": current_title or "토픽",
                        "questions": current_questions,
                    }
                ],
            }
        )
        current_questions = []
        q_in_chapter = 0

    for ri, row in enumerate(rows):
        if ri == 0:
            continue
        item = row[1] if len(row) > 1 else None
        qnum = row[2] if len(row) > 2 else None
        guide = row[3] if len(row) > 3 else None
        note = row[4] if len(row) > 4 else None

        if guide is None or not str(guide).strip():
            continue

        sec = section_from_col_b(item)
        if sec:
            if current_title is not None and sec != current_title and current_questions:
                flush_chapter()
                ch_idx += 1
            elif current_title is None:
                ch_idx = 1
            current_title = sec

        if current_title is None:
            current_title = "인터뷰"
            ch_idx = max(ch_idx, 1)

        raw = str(guide).strip()
        excel_q = str(qnum).strip() if qnum is not None else f"행{ri + 1}"
        probing_guide = f"[{excel_q}]" if excel_q else None

        segment_note, body_after_segment = strip_leading_segment_lines(raw)
        main_q, probes_from_body, info_from_body = split_compound_question(body_after_segment)
        probes_from_body, info_from_body = refine_probes_and_info_from_newlines(
            probes_from_body, info_from_body
        )
        info_from_note, probes_from_note = split_note_into_info_and_probes(note)
        info_all = info_from_note + info_from_body
        checklist_probes = probes_from_body + probes_from_note
        root_note = build_root_note(info_all)
        info_blob = "\n".join(p.strip() for p in info_all if p and str(p).strip())

        q_in_chapter += 1
        readable_qid = f"question_{ch_idx}_{1}_{q_in_chapter}"

        if segment_note:
            question_content = SEGMENT_BRIDGE_CONTENT
            branch_list = build_follow_up_branches(
                question_content,
                info_blob,
                readable_qid,
                MIN_FOLLOW_UP_BRANCHES,
                forced_followup_text=main_q,
                segment_mod_note=segment_note,
                keyword_content=main_q,
            )
        else:
            question_content = main_q
            branch_list = build_follow_up_branches(
                main_q, info_blob, readable_qid, MIN_FOLLOW_UP_BRANCHES
            )

        current_questions.append(
            make_question(
                question_content,
                readable_qid,
                excel_q,
                root_note,
                checklist_probes,
                branch_list,
                probing_guide,
            )
        )

    flush_chapter()

    goal = meta.get(
        "goal",
        "가입 초기 고객 경험 및 NPS 관련 인사이트 수집",
    )
    target = meta.get(
        "target_audience",
        "조사 조건(엑셀 조사설계·조건 시트)에 부합하는 가입 초기 고객",
    )

    title = "LG U+ · 가이드라인 2차 (가입 초기 경험)"
    desc = (
        "LG U+ Poc2 · 엑셀 `가이드라인 2차.xlsx`의 `가이드라인_수정` 시트(항목·번호·인터뷰 가이드·비고) 기반. "
        "생성: build_lg_guideline_2nd_survey_draft.py · 질문별 예상 답변 분기·팔로업: build_follow_up_branches()"
    )

    return {
        "survey_draft": {
            "title": title,
            "description": desc,
            "welcome_title": "안녕하세요 인터뷰에 참여해 주셔서 감사합니다",
            "welcome_message": "오늘 인터뷰는 통신 가입 이후 경험과 인식에 대해 이야기 나눕니다.",
            "goal": goal,
            "target_audience": target,
            "quota": 10,
            "user_groups": [],
            "user_inputs": [
                {"id": nid(), "readable_id": "user_input_1", "label": "이름을 적어 주세요 (또는 닉네임)"}
            ],
            "screener_questions": [],
            "interview_opening": (
                "안녕하세요. 저는 오늘 인터뷰를 진행하는 AI 인터뷰어입니다.\n"
                "정답은 없으며, 경험하신 대로 편하게 말씀해 주시면 됩니다.\n"
                "조건에 해당하지 않는 질문은 건너뛰어도 됩니다.\n"
                "준비되셨으면 '준비되었어요'라고 말씀해 주세요."
            ),
            "interview_chapters": interview_chapters,
            "interview_closing": {
                "closing_message": (
                    "오늘 소중한 시간 내어 주셔서 감사합니다. 추가로 하실 말씀이 있으시면 짧게 말씀해 주시고, "
                    "없으시면 '없습니다'라고 해 주셔도 됩니다. 감사합니다."
                )
            },
        }
    }


def fix_typos(obj):
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "content" and isinstance(v, str):
                obj[k] = v.replace("무엇기 가장", "무엇이 가장").replace("무엇기 ", "무엇이 ")
            else:
                fix_typos(v)
    elif isinstance(obj, list):
        for x in obj:
            fix_typos(x)


def main():
    data = build_from_excel()
    fix_typos(data)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    nq = sum(len(t["questions"]) for ch in data["survey_draft"]["interview_chapters"] for t in ch["topics"])
    print(f"Wrote {OUT_JSON} ({nq} questions, {len(data['survey_draft']['interview_chapters'])} chapters)")


if __name__ == "__main__":
    main()
